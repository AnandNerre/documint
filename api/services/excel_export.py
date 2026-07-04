"""Generate Excel workbook from extracted document fields."""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from schemas import SalarySlipFields

_LABELS: list[tuple[str, str]] = [
    ("employee_name", "Employee Name"),
    ("employee_id", "Employee ID"),
    ("company_name", "Company"),
    ("designation", "Designation"),
    ("pay_period", "Pay Period"),
    ("currency", "Currency"),
    ("basic_salary", "Basic Salary"),
    ("hra", "Housing Allowance"),
    ("gross_salary", "Gross Salary"),
    ("total_deductions", "Total Deductions"),
    ("net_pay", "Net Pay"),
    ("confidence_notes", "Notes"),
]

_CURRENCY_FIELDS = {"basic_salary", "hra", "gross_salary", "total_deductions", "net_pay"}

_CURRENCY_SYMBOLS = {
    "INR": "₹",
    "USD": "$",
    "EUR": "€",
    "GBP": "£",
    "AED": "AED ",
    "SAR": "SAR ",
    "SGD": "S$",
    "CAD": "C$",
    "AUD": "A$",
    "MYR": "RM",
    "PHP": "₱",
    "ZAR": "R",
    "NGN": "₦",
}


def _money_format(currency: str | None) -> str:
    code = (currency or "USD").upper()
    sym = _CURRENCY_SYMBOLS.get(code, f"{code} ")
    return f'"{sym}"#,##0.00'


def build_workbook(fields: SalarySlipFields) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    header_fill = PatternFill("solid", fgColor="1E8E3E")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    label_font = Font(bold=True, size=11)
    thin = Side(style="thin", color="DADCE0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = _money_format(fields.currency)

    ws.merge_cells("A1:B1")
    title = ws["A1"]
    title.value = "DocuMint — Export"
    title.font = header_font
    title.fill = header_fill
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Exported"
    ws["B2"] = datetime.now().strftime("%d %b %Y, %I:%M %p")
    ws["A2"].font = Font(color="5F6368")
    ws["B2"].font = Font(color="5F6368")

    row = 4
    for key, label in _LABELS:
        value = getattr(fields, key)
        ws.cell(row=row, column=1, value=label).font = label_font
        cell = ws.cell(row=row, column=2)

        if key in _CURRENCY_FIELDS and value is not None:
            cell.value = float(value)
            cell.number_format = money_fmt
        else:
            cell.value = value if value is not None else ""

        for col in (1, 2):
            c = ws.cell(row=row, column=col)
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 36

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
